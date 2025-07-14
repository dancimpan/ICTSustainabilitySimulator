# utils.py

"""
Acest modul oferă funcții utilitare pentru simulator, în special pentru
calculele de conversie de la costuri abstracte la impactul din lumea reală
și pentru generarea de rapoarte.
"""
import io
import pandas as pd
from openpyxl.drawing.image import Image

# ==============================================================================
# == SECȚIUNEA 1: CONSTANTE PENTRU ECHIVALENTE REALE
# ==============================================================================
# Sursa pentru eficiență EV: Medie generală.
EV_EFFICIENCY_KWH_PER_KM = 0.18
# Sursa pentru absorbție copac: Un copac matur absoarbe ~22kg CO2/an.
# Calcul: 22000 grame / 365 zile / 24 ore
GCO2_ABSORBED_BY_TREE_PER_HOUR = 22000.0 / 365.0 / 24.0


# ==============================================================================
# == SECȚIUNEA 2: FUNCȚII DE CALCUL PRINCIPALE
# ==============================================================================

def calculate_energy_co2(cpu_ops, data_movement, kwh_cpu_factor, kwh_data_factor, gco2eq_per_kwh_factor):
    """
    Calculează energia totală estimată (kWh) și emisiile de CO2 (g)
    pe baza costurilor abstracte și a factorilor de conversie.
    """
    energy_from_cpu = cpu_ops * kwh_cpu_factor
    energy_from_data = data_movement * kwh_data_factor
    estimated_kwh = energy_from_cpu + energy_from_data
    estimated_co2_g = estimated_kwh * gco2eq_per_kwh_factor
    return estimated_kwh, estimated_co2_g


def get_real_world_equivalents(estimated_co2_g, gco2eq_per_kwh_factor):
    """
    Convertește o cantitate dată de emisii de CO2 (în grame) în echivalente
    tangibile, din lumea reală.
    """
    equivalents = {}
    if estimated_co2_g > 0.0001:
        # Calculăm emisiile de gCO2 per km pentru o mașină electrică
        # pe baza eficienței (kWh/km) și a factorului de emisii al rețelei (gCO2/kWh)
        gco2_per_km_ev = EV_EFFICIENCY_KWH_PER_KM * gco2eq_per_kwh_factor
        
        if gco2_per_km_ev > 0:
            km_driven_ev = estimated_co2_g / gco2_per_km_ev
            equivalents["km parcurși cu o mașină electrică"] = km_driven_ev
            
        tree_hours = estimated_co2_g / GCO2_ABSORBED_BY_TREE_PER_HOUR
        equivalents["ore necesare unui copac pentru a absorbi"] = tree_hours
        
    return equivalents

# ==============================================================================
# == SECȚIUNEA 3: FUNCȚIE PENTRU EXPORT EXCEL
# ==============================================================================

def create_excel_export(all_results, df_reductions, df_history, figs_cost, figs_impact):
    """
    Creează un fișier Excel în memorie, conținând toate datele și graficele unei simulări.

    Args:
        all_results (list): Lista de dicționare cu rezultatele detaliate pentru fiecare model.
        df_reductions (pd.DataFrame): DataFrame-ul cu reducerile procentuale.
        df_history (pd.DataFrame): DataFrame-ul cu istoricul complet al rulărilor.
        figs_cost (dict): Un dicționar de figuri Plotly pentru graficele de costuri.
        figs_impact (dict): Un dicționar de figuri Plotly pentru graficele de impact.

    Returns:
        bytes: Conținutul binar al fișierului Excel generat.
    """
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Foaia 1: Rezumat și Reduceri
        df_summary = pd.DataFrame(all_results)
        df_summary.to_excel(writer, sheet_name='Rezumat Simulare', index=False, startrow=0)
        if not df_reductions.empty:
            df_reductions.to_excel(writer, sheet_name='Rezumat Simulare', index=True, startrow=len(df_summary) + 3)

        # Foaia 2: Istoricul Complet
        if not df_history.empty:
            df_history.to_excel(writer, sheet_name='Istoric Complet', index=False)

        # Funcție ajutătoare pentru a adăuga grafice
        def add_charts_to_sheet(sheet_name, figs_dict):
            if sheet_name not in writer.book.sheetnames:
                 ws = writer.book.create_sheet(sheet_name)
            else:
                 ws = writer.book[sheet_name]
                 
            current_row = 1
            for title, fig in figs_dict.items():
                img_data = io.BytesIO(fig.to_image(format="png", width=800, height=450, scale=2))
                img = Image(img_data)
                
                ws.cell(row=current_row, column=1, value=title)
                ws.add_image(img, f'A{current_row + 1}')
                current_row += 25 

        # Adăugăm foile de calcul cu grafice
        if figs_impact:
            add_charts_to_sheet('Grafice Impact Ambiental', figs_impact)
        if figs_cost:
            add_charts_to_sheet('Grafice Costuri Abstracte', figs_cost)

    processed_data = output.getvalue()
    return processed_data
